import json
import os
import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

DATA_FILE = "data.json"
EXCEL_FILE = "airbnb_data.xlsx"

COMISIONES = {"Airbnb": 0.15, "Booking": 0.12, "Directo": 0.03}
DIAS_POR_MES = {
    "Enero":31,"Febrero":28,"Marzo":31,"Abril":30,"Mayo":31,"Junio":30,
    "Julio":31,"Agosto":31,"Septiembre":30,"Octubre":31,"Noviembre":30,"Diciembre":31
}
MESES_ES = {1:"Enero",2:"Febrero",3:"Marzo",4:"Abril",5:"Mayo",6:"Junio",
            7:"Julio",8:"Agosto",9:"Septiembre",10:"Octubre",11:"Noviembre",12:"Diciembre"}

class DataManager:
    def __init__(self):
        self._cargar_datos()

    def _cargar_datos(self):
        if os.path.exists(DATA_FILE):
            with open(DATA_FILE, "r") as f:
                self.datos = json.load(f)
        else:
            self.datos = {"reservas": [], "gastos": [], "config": {"inversion": 25000}}

    def _guardar_datos(self):
        with open(DATA_FILE, "w") as f:
            json.dump(self.datos, f, indent=2, ensure_ascii=False)

    def _parsear_fecha(self, fecha_str):
        """Acepta DD/MM o DD/MM/YYYY"""
        partes = fecha_str.split("/")
        dia, mes = int(partes[0]), int(partes[1])
        año = int(partes[2]) if len(partes) > 2 else datetime.datetime.now().year
        return datetime.date(año, mes, dia)

    def agregar_reserva(self, entrada_str, salida_str, plataforma, tarifa, huesped=""):
        entrada = self._parsear_fecha(entrada_str)
        salida  = self._parsear_fecha(salida_str)
        noches  = (salida - entrada).days
        bruto   = noches * tarifa
        com_pct = COMISIONES.get(plataforma, 0.15)
        comision = bruto * com_pct
        neto    = bruto - comision

        reserva = {
            "id": len(self.datos["reservas"]) + 1,
            "entrada": entrada.strftime("%d/%m/%Y"),
            "salida": salida.strftime("%d/%m/%Y"),
            "mes": MESES_ES[entrada.month],
            "año": entrada.year,
            "plataforma": plataforma,
            "huesped": huesped,
            "tarifa": tarifa,
            "noches": noches,
            "bruto": round(bruto, 2),
            "comision": round(comision, 2),
            "com_pct": int(com_pct * 100),
            "neto": round(neto, 2),
        }
        self.datos["reservas"].append(reserva)
        self._guardar_datos()
        return reserva

    def agregar_gasto(self, categoria, descripcion, monto):
        hoy = datetime.date.today()
        
        # Calcular total del mes actual
        mes_actual = MESES_ES[hoy.month]
        total_mes = sum(
            g["monto"] for g in self.datos["gastos"]
            if g["mes"] == mes_actual and g["año"] == hoy.year
        ) + monto

        gasto = {
            "id": len(self.datos["gastos"]) + 1,
            "fecha": hoy.strftime("%d/%m/%Y"),
            "mes": mes_actual,
            "año": hoy.year,
            "categoria": categoria,
            "descripcion": descripcion,
            "monto": round(monto, 2),
            "tipo": "Fijo" if categoria in ["Servicios","Seguro","Impuestos"] else "Variable",
        }
        self.datos["gastos"].append(gasto)
        self._guardar_datos()
        return {"fecha": gasto["fecha"], "total_mes": round(total_mes, 2)}

    def resumen_mes(self, mes_nombre):
        año = datetime.datetime.now().year
        reservas = [r for r in self.datos["reservas"] if r["mes"] == mes_nombre and r["año"] == año]
        gastos   = [g for g in self.datos["gastos"]   if g["mes"] == mes_nombre and g["año"] == año]

        noches   = sum(r["noches"] for r in reservas)
        bruto    = sum(r["bruto"] for r in reservas)
        comisiones = sum(r["comision"] for r in reservas)
        neto     = sum(r["neto"] for r in reservas)
        fijos    = sum(g["monto"] for g in gastos if g["tipo"] == "Fijo")
        variables = sum(g["monto"] for g in gastos if g["tipo"] == "Variable")
        gastos_t = fijos + variables
        utilidad = neto - gastos_t
        dias_mes = DIAS_POR_MES.get(mes_nombre, 30)

        return {
            "mes": mes_nombre,
            "noches": noches,
            "num_reservas": len(reservas),
            "ocupacion": noches / dias_mes if dias_mes > 0 else 0,
            "bruto": bruto,
            "comisiones": comisiones,
            "neto": neto,
            "fijos": fijos,
            "variables": variables,
            "gastos": gastos_t,
            "utilidad": utilidad,
            "margen": utilidad / bruto if bruto > 0 else 0,
            "adr": bruto / noches if noches > 0 else 0,
        }

    def kpis_anuales(self):
        año = datetime.datetime.now().year
        reservas = [r for r in self.datos["reservas"] if r["año"] == año]
        gastos   = [g for g in self.datos["gastos"]   if g["año"] == año]

        noches = sum(r["noches"] for r in reservas)
        neto   = sum(r["neto"] for r in reservas)
        bruto  = sum(r["bruto"] for r in reservas)
        gastos_t = sum(g["monto"] for g in gastos)
        utilidad = neto - gastos_t
        inversion = self.datos["config"].get("inversion", 25000)

        return {
            "noches_total": noches,
            "ocupacion": noches / 365,
            "adr": bruto / noches if noches > 0 else 0,
            "revpar": bruto / 365,
            "ingreso_neto": neto,
            "gastos": gastos_t,
            "utilidad": utilidad,
            "roi": utilidad / inversion if inversion > 0 else 0,
            "payback": inversion / (utilidad / 12) if utilidad > 0 else 999,
        }

    def ultimas_entradas(self, n=5):
        entradas = []
        for r in self.datos["reservas"][-n:]:
            entradas.append({
                "tipo": "reserva",
                "resumen": f"{r['entrada']} → {r['salida']} · {r['plataforma']} · ${r['neto']:.0f} neto"
            })
        for g in self.datos["gastos"][-(n - len(entradas)):]:
            entradas.append({
                "tipo": "gasto",
                "resumen": f"{g['fecha']} · {g['categoria']} · ${g['monto']:.0f}"
            })
        return entradas[-n:]

    def exportar_excel(self):
        wb = Workbook()
        
        # ── Hoja Reservas ──
        ws1 = wb.active
        ws1.title = "Reservas"
        hdrs = ["#","Entrada","Salida","Noches","Mes","Plataforma","Huesped","Tarifa/Noche","Bruto","Comision","Neto"]
        for i,h in enumerate(hdrs,1):
            c = ws1.cell(1,i,h)
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="0D1B2A")
        
        for i,r in enumerate(self.datos["reservas"],2):
            vals = [r["id"],r["entrada"],r["salida"],r["noches"],r["mes"],
                    r["plataforma"],r["huesped"],r["tarifa"],r["bruto"],r["comision"],r["neto"]]
            for j,v in enumerate(vals,1):
                ws1.cell(i,j,v)
        
        # ── Hoja Gastos ──
        ws2 = wb.create_sheet("Gastos")
        hdrs2 = ["#","Fecha","Mes","Categoria","Tipo","Descripcion","Monto"]
        for i,h in enumerate(hdrs2,1):
            c = ws2.cell(1,i,h)
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="B71C1C")
        
        for i,g in enumerate(self.datos["gastos"],2):
            vals = [g["id"],g["fecha"],g["mes"],g["categoria"],g["tipo"],g["descripcion"],g["monto"]]
            for j,v in enumerate(vals,1):
                ws2.cell(i,j,v)
        
        # ── Hoja Resumen Mensual ──
        ws3 = wb.create_sheet("Resumen Mensual")
        meses = list(DIAS_POR_MES.keys())
        cols = ["Mes","Noches","Reservas","Bruto","Comisiones","Neto","Gastos","Utilidad","Ocupacion%","ADR"]
        for i,h in enumerate(cols,1):
            c = ws3.cell(1,i,h)
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="1B5E20")
        
        for i,mes in enumerate(meses,2):
            rm = self.resumen_mes(mes)
            vals = [mes, rm["noches"], rm["num_reservas"], rm["bruto"],
                    rm["comisiones"], rm["neto"], rm["gastos"], rm["utilidad"],
                    f"{rm['ocupacion']*100:.1f}%", f"${rm['adr']:.2f}"]
            for j,v in enumerate(vals,1):
                ws3.cell(i,j,v)
        
        wb.save(EXCEL_FILE)
        return EXCEL_FILE
